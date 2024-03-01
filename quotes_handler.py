import datetime as dt
from nsetools import Nse
from nselib import capital_market
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
import requests
from AppCore import constants


class QuotesHandler:
    index_list = []
    quotes_list = []
    nse = None

    def __init__(self):

        self.nse = Nse()

    def get_advance_decline_index_list(self, percentage):
        """
        :returns a INDEX name, in which the percentage advances are
        equal or greater than that percentage mentioned
        :param percentage:
        """
        adv_dec_df = pd.DataFrame(self.nse.get_advances_declines())
        date_string = datetime.now().strftime(constants.DATE_INDIAN_FORMAT_STRING)
        excel_file = constants.INDEX_DATA_FILES_DIR_PATH + f"{date_string}_advdec.xlsx"
        adv_dec_df.to_excel(excel_file, index=False)

        percentage = (percentage / 100.0)
        for index, indexinfo in adv_dec_df.iterrows():
            adv = indexinfo["advances"]
            dec = indexinfo["declines"]
            unchanged = indexinfo["unchanged"]

            total = (adv + dec + unchanged)
            lper = adv / total
            if lper > percentage:
                self.index_list.append(indexinfo["indice"])
                return self.index_list

    def get_sectoral_quote_information(self,sectoral_code):
        print(self.nse.get_index_quote(sectoral_code, as_json=False))

    def get_BHAV_copy(self, date=datetime.now()):

        date_string = date.strftime(constants.DATE_INDIAN_FORMAT_STRING)
        excel_file = constants.BHAV_FILES_DIR_PATH + f"{date_string}_bhav.xlsx"
        delivery = pd.DataFrame(capital_market.bhav_copy_equities(date.strftime(constants.DATE_INDIAN_FORMAT_STRING)))
        delivery.to_excel(excel_file, index=False)

    def get_BHAV_copy_WithDelivery(self, date=datetime.now()):

        date_string = date.strftime(constants.DATE_INDIAN_FORMAT_STRING)
        excel_file = constants.BHAV_DELIVERY_FILES_DIR_PATH + f"{date_string}_delivery.xlsx"
        delivery = pd.DataFrame(
            capital_market.bhav_copy_with_delivery(date.strftime(constants.DATE_INDIAN_FORMAT_STRING)))
        delivery.to_excel(excel_file, index=False)

    def get_quotes_list(self):
        equity_quotes_df = capital_market.equity_list()
        equity_quotes_df.to_excel(constants.QUOTE_FILES_DIR_PATH + 'equitySymbols.xlsx', index=False)

        # Load the Excel file
        wb = load_workbook(filename=constants.QUOTE_FILES_DIR_PATH + 'equitySymbols.xlsx')
        # Select the worksheet
        ws = wb['Sheet1']
        # Access cell values
        # Iterate over rows
        for row in ws.iter_rows(values_only=True):
            if row[2] == 'EQ':
                self.quotes_list.append(row[0])
        # Return List
        return self.quotes_list

    def get_index_quotes_list(self, index='NIFTY 250', from_web=False):
        """
        get list of all equity available to trade in NSE
        Valid values
        index = NIFTY 250
        index = NIFTY 500
        :return: pandas data frame
        """
        data_df = pd.DataFrame(index=['NAME OF COMPANY', 'INDUSTRY', 'SYMBOL', ' SERIES', 'ISIN CODE'])

        if from_web:
            try:
                # URL of the CSV file
                url250 = 'https://nsearchives.nseindia.com/content/indices/ind_niftysmallcap250list.csv'
                url500 = "https://nsearchives.nseindia.com/content/indices/ind_nifty500list.csv"
                response = None
                # Fetch the CSV file from the URL
                if index == 'NIFTY 500':
                    response = requests.get(url500)
                elif index == 'NIFTY 250':
                    response = requests.get(url250)

                # Check if the request was successful
                if response.status_code == 200:
                    data_df = pd.read_csv(url500)
                else:
                    print("Failed to fetch the CSV file.")

            except Exception as e:
                raise FileNotFoundError(f' Equity List not found :: NSE error : {e}')
        else:
            if index == 'NIFTY 500':
                data_df = pd.read_csv(constants.QUOTE_FILES_DIR_PATH + 'ind_nifty500list.csv', index_col=None)
            elif index == 'NIFTY 250':
                data_df = pd.read_csv(constants.QUOTE_FILES_DIR_PATH + 'ind_niftysmallcap250list.csv', index_col=None)

        return data_df['Symbol']

    def save_selected_quotes_to_excel(self, selected_quotes_list):
        # Create a new Excel workbook
        wb = Workbook()
        # Select the active worksheet
        ws = wb.active
        # Write data to cells
        ws['A1'] = 'Quote Short Name'
        for rowIndex, colValue in enumerate(selected_quotes_list):
            cell = ws.cell(row=rowIndex + 2, column=1)
            value = selected_quotes_list[rowIndex]
            cell.value = value
        # Save the workbook
        wb.save(constants.QUOTE_FILES_DIR_PATH + 'ShortListed.xlsx')

    def get_block_bulk_deal_data(self, period):

        block_deals_df = capital_market.block_deals_data(period='1Y')
        block_deal_file_str = constants.BLOCK_DEAL_FILES_DIR_PATH + f'{datetime.now().strftime(constants.DATE_INDIAN_FORMAT_STRING)}_block.xlsx'
        block_deals_df.to_excel(block_deal_file_str, index=False)

        bulk_deal_file_str = constants.BULK_DEALS_FILES_DIR_PATH + f'{datetime.now().strftime(constants.DATE_INDIAN_FORMAT_STRING)}_bulk.xlsx'
        bulk_deals_df = capital_market.bulk_deal_data(period='1Y')
        bulk_deals_df.to_excel(bulk_deal_file_str, index=False)

    def get_market_watch(self):

        bulk_deal_file_str = constants.MW_DEALS_FILES_DIR_PATH + f'{datetime.now().strftime(constants.DATE_INDIAN_FORMAT_STRING)}_watch.xlsx'
        market_watch_df = capital_market.market_watch_all_indices()
        market_watch_df.to_excel(bulk_deal_file_str, index=False)
