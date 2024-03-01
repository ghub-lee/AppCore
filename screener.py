from AppCore import quotes_handler
from datetime import datetime, timedelta
from openpyxl import Workbook
import pandas as pd
import constants

qh_instance = quotes_handler.QuotesHandler()


class Screener:
    screened_list = []
    static_variable = 2
    fyers = None

    def __init__(self):
        pass

    def get_stock_price_ranged_previous_day(self, minimum_price, estimated_range, from_web=False):

        shortlisted_with_price = []
        qh_instance.get_BHAV_copy(datetime.now()-timedelta(days=1))
        quotes_list = qh_instance.get_index_quotes_list()
        previous_date = datetime.now() - timedelta(days=1)
        previous_date_str = previous_date.strftime(constants.DATE_INDIAN_FORMAT_STRING)
        df = pd.read_excel(constants.BHAV_FILES_DIR_PATH + f'{previous_date_str}_bhav.xlsx')
        print(df.head())


        wb = Workbook()
        # Select the active worksheet
        ws = wb.active
        # Write data to cells
        ws['A1'] = 'Short Name'
        ws['B1'] = 'LTP'

        for rowIndex, quote_short_name in enumerate(quotes_list):
            ltp = df[df['SYMBOL'] == quote_short_name]['CLOSE']
            pltp = ltp.array[0]

            # using loc accessor
            # pltp = df.loc[df['SYMBOL'] == quote_short_name, 'CLOSE']
            print(quote_short_name, pltp)
            if from_web:
                data = {"symbols": "NSE:" + quote_short_name + "-EQ"}
                response = self.fyers.quotes(data)
                if response["code"] == 200:
                    # This should be changed if you provide
                    # multiple script names index:0,1,2,..
                    dt = response["d"][0]["v"]
                    pltp = dt["prev_close_price"]
                    if minimum_price + estimated_range > pltp > minimum_price:
                        self.screened_list.append(quote_short_name)
                        short_name_cell = ws.cell(row=self.static_variable, column=1)
                        short_prvprice_cell = ws.cell(row=self.static_variable, column=2)
                        short_name_cell.value = quote_short_name
                        short_prvprice_cell.value = pltp
                        self.static_variable += 1
                        print("Screening Completed", quote_short_name, pltp)
            else:
                if minimum_price + estimated_range > pltp > minimum_price:
                    shortlisted_with_price.append(quote_short_name)
        return shortlisted_with_price


        # Save the workbook
        # wb.save('./Data/QuoteFile/ShortListedWithScreener.xlsx')

    def viewHistoricalData(self):

        to_time_string = "2024-02-19 23:59:00"
        from_time_string = "2024-02-19 00:00:00"

        # Get the current date and time
        current_date = datetime.now()

        # Calculate the difference for 100 days
        difference = timedelta(days=100)

        ftime_object = datetime.strptime(from_time_string, "%Y-%m-%d %H:%M:%S")
        fepochtime = int(ftime_object.timestamp())
        print(fepochtime)

        ttime_object = datetime.strptime(to_time_string, "%Y-%m-%d %H:%M:%S")
        tepochtime = int(ttime_object.timestamp())
        print(tepochtime)

        # fepoch_time = current_date - difference

        # Historical Data
        data = {"symbol": "NSE:SBIN-EQ", "resolution": "10", "date_format": "0", "range_from": str(fepochtime),
                "range_to": str(tepochtime), "cont_flag": "1"}

        hdata = self.fyers.history(data)

        # Create a new Excel workbook
        wb = Workbook()
        # Select the active worksheet
        ws = wb.active
        # Write data to cells
        ws['A1'] = 'Date'
        ws['B1'] = 'Open'
        ws['C1'] = 'High'
        ws['D1'] = 'Low'
        ws['E1'] = 'Close'
        ws['F1'] = 'Volume'

        # Create a pandas DataFrame from the data
        df = pd.DataFrame(hdata)
        # Access the first column using indexing
        first_column = df.iloc[:, 0]
        for rowIndex, rowValue in enumerate(first_column):
            for colIndex, colValue in enumerate(rowValue):
                cell = ws.cell(row=rowIndex + 2, column=colIndex + 1)
                value = rowValue[colIndex]
                if colIndex == 0:
                    value = rowValue[colIndex]
                    # Convert epoch time to a datetime object
                    date_time = datetime.utcfromtimestamp(value)
                    # Convert datetime object to a string with a specific format
                    formatted_date_time = date_time.strftime('%Y-%m-%d %H:%M:%S')
                    # Updating Cell Value
                    cell.value = formatted_date_time
                else:
                    # Updating Cell Value
                    cell.value = value

        # Save the workbook
        wb.save('output.xlsx')
        print("Data exported")
